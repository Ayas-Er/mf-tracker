"""
Mutual Fund Investment Tracker
Streamlit Cloud app with Google Sheets storage
"""

import streamlit as st
import pandas as pd
import numpy as np
import requests
import gspread
from google.oauth2.service_account import Credentials
from datetime import datetime, date
from scipy.optimize import newton, brentq
import io
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import uuid

st.set_page_config(
    page_title="MF Investment Tracker",
    page_icon="📈",
    layout="wide"
)

# ─── GOOGLE SHEETS SETUP ────────────────────────────────────────────────────
SPREADSHEET_URL = "https://docs.google.com/spreadsheets/d/1TPlxbemlHM2DGZRxHbiQlE-xZmuNokGWEBMpO8t9zGs"
SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive"
]

SHEET_FUNDS  = "Funds"
SHEET_TRADES = "Trades"

FUNDS_COLS = [
    "fund_id", "amfi_code", "fund_name",
    "target_pct", "reinvest_pct", "reinvest_months", "added_on"
]

TRADES_COLS = [
    "trade_id", "fund_id", "fund_name", "buy_date", "buy_nav",
    "units", "remaining_units", "invested_amount", "sip_amount",
    "reinvest_amount", "target_nav", "status",
    "sell_date", "sell_nav", "sell_value", "gain",
    "holding_days", "tax_type"
]

@st.cache_resource
def get_gspread_client():
    creds_dict = dict(st.secrets["gcp_service_account"])
    creds = Credentials.from_service_account_info(creds_dict, scopes=SCOPES)
    return gspread.authorize(creds)

def get_sheet(sheet_name):
    client = get_gspread_client()
    sh = client.open_by_url(SPREADSHEET_URL)
    try:
        return sh.worksheet(sheet_name)
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet(title=sheet_name, rows=1000, cols=30)
        if sheet_name == SHEET_FUNDS:
            ws.append_row(FUNDS_COLS)
        elif sheet_name == SHEET_TRADES:
            ws.append_row(TRADES_COLS)
        return ws

def read_sheet(sheet_name, expected_cols):
    ws = get_sheet(sheet_name)
    data = ws.get_all_records(expected_headers=expected_cols)
    if not data:
        return pd.DataFrame(columns=expected_cols)
    df = pd.DataFrame(data)
    for col in expected_cols:
        if col not in df.columns:
            df[col] = ""
    return df[expected_cols]

def append_row(sheet_name, row_dict, expected_cols):
    ws = get_sheet(sheet_name)
    row = [str(row_dict.get(c, "")) for c in expected_cols]
    ws.append_row(row)

def update_trade_row(trade_id, updates: dict):
    ws = get_sheet(SHEET_TRADES)
    records = ws.get_all_records(expected_headers=TRADES_COLS)
    for i, rec in enumerate(records):
        if rec["trade_id"] == trade_id:
            row_num = i + 2  # 1-indexed + header
            for col_name, val in updates.items():
                if col_name in TRADES_COLS:
                    col_idx = TRADES_COLS.index(col_name) + 1
                    ws.update_cell(row_num, col_idx, str(val))
            break

# ─── NAV FETCH ───────────────────────────────────────────────────────────────
@st.cache_data(ttl=3600)
def fetch_nav(amfi_code, for_date=None):
    try:
        url  = f"https://api.mfapi.in/mf/{amfi_code}"
        resp = requests.get(url, timeout=10).json()
        name = resp["meta"]["scheme_name"]

        if for_date:
            # Find NAV for the given date or nearest previous date
            nav_data = pd.DataFrame(resp["data"])
            nav_data["date"] = pd.to_datetime(nav_data["date"], format="%d-%m-%Y")
            nav_data["nav"]  = nav_data["nav"].astype(float)
            nav_data = nav_data.sort_values("date")

            target_date = pd.Timestamp(for_date)
            # Get NAV on or before the trade date (markets may be closed)
            available = nav_data[nav_data["date"] <= target_date]
            if available.empty:
                return name, None, None
            row      = available.iloc[-1]
            nav      = row["nav"]
            nav_date = row["date"].strftime("%d-%m-%Y")
        else:
            nav      = float(resp["data"][0]["nav"])
            nav_date = resp["data"][0]["date"]

        return name, nav, nav_date
    except Exception:
        return None, None, None

@st.cache_data(ttl=86400)
def search_funds(query):
    try:
        url  = "https://api.mfapi.in/mf/search"
        resp = requests.get(url, params={"q": query}, timeout=10).json()
        return pd.DataFrame(resp) if resp else pd.DataFrame()
    except Exception:
        return pd.DataFrame()

# ─── XIRR ────────────────────────────────────────────────────────────────────
def xirr(cashflows, dates):
    if len(cashflows) < 2:
        return 0.0
    dates = [pd.Timestamp(d) for d in dates]
    def xnpv(rate):
        if rate <= -1:
            return float('inf')
        return sum(
            cf / ((1 + rate) ** min((d - dates[0]).days / 365, 300))
            for cf, d in zip(cashflows, dates)
        )
    for guess in [0.10, 0.01, 0.5, -0.1, 0.3]:
        try:
            r = newton(xnpv, guess, maxiter=1000, tol=1e-6)
            if -1 < r < 100:
                return r
        except RuntimeError:
            continue
    try:
        return brentq(xnpv, -0.999, 100, maxiter=1000)
    except ValueError:
        return 0.0

# ─── FIFO SELL ────────────────────────────────────────────────────────────────
def process_fifo_sell(fund_id, sell_date, sell_nav, units_to_sell):
    trades_df = read_sheet(SHEET_TRADES, TRADES_COLS)
    if trades_df.empty:
        return [], "No trades found."

    trades_df["remaining_units"] = pd.to_numeric(trades_df["remaining_units"], errors="coerce").fillna(0)
    trades_df["buy_nav"]         = pd.to_numeric(trades_df["buy_nav"],         errors="coerce").fillna(0)
    trades_df["target_nav"]      = pd.to_numeric(trades_df["target_nav"],      errors="coerce").fillna(0)
    trades_df["buy_date"]        = pd.to_datetime(trades_df["buy_date"],        errors="coerce")

    open_lots = (
        trades_df[
            (trades_df["fund_id"] == fund_id) &
            (trades_df["status"] == "Open") &
            (trades_df["remaining_units"] > 0)
        ]
        .sort_values("buy_date")
        .copy()
    )

    if open_lots.empty:
        return [], "No open lots for this fund."

    remaining    = float(units_to_sell)
    sell_records = []

    for _, lot in open_lots.iterrows():
        if remaining <= 0:
            break
        available    = float(lot["remaining_units"])
        sell_units   = min(available, remaining)
        sell_value   = sell_units * float(sell_nav)
        buy_value    = sell_units * float(lot["buy_nav"])
        gain         = sell_value - buy_value
        holding_days = (pd.Timestamp(sell_date) - pd.Timestamp(lot["buy_date"])).days
        tax_type     = "LTCG" if holding_days >= 365 else "STCG"
        new_remaining = available - sell_units
        new_status    = "Closed" if new_remaining < 0.0001 else "Open"

        sell_records.append({
            "trade_id":      lot["trade_id"],
            "buy_date":      lot["buy_date"],
            "buy_nav":       lot["buy_nav"],
            "sell_units":    sell_units,
            "sell_value":    sell_value,
            "buy_value":     buy_value,
            "gain":          gain,
            "holding_days":  holding_days,
            "tax_type":      tax_type,
            "new_remaining": new_remaining,
            "new_status":    new_status,
        })
        remaining -= sell_units

    return sell_records, None

def commit_fifo_sell(sell_records, sell_date, sell_nav):
    for rec in sell_records:
        update_trade_row(rec["trade_id"], {
            "remaining_units": round(rec["new_remaining"], 6),
            "status":          rec["new_status"],
            "sell_date":       str(sell_date),
            "sell_nav":        round(float(sell_nav), 4),
            "sell_value":      round(rec["sell_value"], 2),
            "gain":            round(rec["gain"], 2),
            "holding_days":    rec["holding_days"],
            "tax_type":        rec["tax_type"],
        })

# ─── ANALYTICS ───────────────────────────────────────────────────────────────
def compute_analytics(fund_id=None):
    trades_df = read_sheet(SHEET_TRADES, TRADES_COLS)
    funds_df  = read_sheet(SHEET_FUNDS,  FUNDS_COLS)

    if trades_df.empty:
        return {}

    num_cols = ["buy_nav", "units", "remaining_units", "invested_amount",
                "sip_amount", "reinvest_amount", "target_nav",
                "sell_nav", "sell_value", "gain", "holding_days"]
    for col in num_cols:
        trades_df[col] = pd.to_numeric(trades_df[col], errors="coerce").fillna(0)

    trades_df["buy_date"]  = pd.to_datetime(trades_df["buy_date"],  errors="coerce")
    trades_df["sell_date"] = pd.to_datetime(trades_df["sell_date"], errors="coerce")

    if fund_id:
        trades_df = trades_df[trades_df["fund_id"] == fund_id]

    closed = trades_df[trades_df["status"] == "Closed"].copy()
    open_t = trades_df[trades_df["status"] == "Open"].copy()

    # ── Tax ──────────────────────────────────────────────────────────────
    tax_df = pd.DataFrame()
    if not closed.empty:
        closed["FY"] = closed["sell_date"].apply(
            lambda d: d.year if pd.notna(d) and d.month > 3 else (d.year - 1 if pd.notna(d) else 0)
        )
        closed["STCG Gain"] = np.where(closed["tax_type"] == "STCG", closed["gain"], 0)
        closed["LTCG Gain"] = np.where(closed["tax_type"] == "LTCG", closed["gain"], 0)
        tax_df = closed.groupby("FY", as_index=False)[["STCG Gain", "LTCG Gain"]].sum()
        tax_df["STCG Tax"] = tax_df["STCG Gain"].clip(lower=0) * 0.20
        tax_df["LTCG Tax"] = (tax_df["LTCG Gain"] - 125000).clip(lower=0) * 0.125
        tax_df["Tax Paid"] = tax_df["STCG Tax"] + tax_df["LTCG Tax"]

    # ── XIRR ─────────────────────────────────────────────────────────────
    cashflows, cf_dates = [], []
    for _, row in trades_df.iterrows():
        cashflows.append(-(row["sip_amount"] + row["reinvest_amount"]))
        cf_dates.append(row["buy_date"])
    if not closed.empty:
        for _, row in closed.iterrows():
            cashflows.append(row["sell_value"])
            cf_dates.append(row["sell_date"])

    terminal = 0
    if not open_t.empty and not funds_df.empty:
        for fid in open_t["fund_id"].unique():
            fund_row = funds_df[funds_df["fund_id"] == fid]
            if fund_row.empty:
                continue
            _, nav, _ = fetch_nav(fund_row.iloc[0]["amfi_code"])
            if nav:
                terminal += open_t[open_t["fund_id"] == fid]["remaining_units"].sum() * nav

    if terminal > 0:
        cashflows.append(terminal)
        cf_dates.append(pd.Timestamp.today())

    pre_tax_xirr = xirr(cashflows, cf_dates)

    cf_after, dt_after = cashflows.copy(), cf_dates.copy()
    if not tax_df.empty:
        for _, row in tax_df.iterrows():
            if row["Tax Paid"] > 0:
                cf_after.append(-row["Tax Paid"])
                dt_after.append(pd.Timestamp(f"{int(row['FY']) + 1}-03-31"))
    after_tax_xirr = xirr(cf_after, dt_after)

    # ── Monthly Summary ───────────────────────────────────────────────────
    monthly_rows = []
    if not trades_df.empty:
        all_months = set(trades_df["buy_date"].dt.to_period("M").dropna())
        if not closed.empty:
            all_months |= set(closed["sell_date"].dt.to_period("M").dropna())
        for p in sorted(all_months):
            mask_buy  = trades_df["buy_date"].dt.to_period("M") == p
            mask_sell = closed["sell_date"].dt.to_period("M") == p if not closed.empty else pd.Series(False, index=closed.index)
            sip      = trades_df.loc[mask_buy, "sip_amount"].sum()
            reinvest = trades_df.loc[mask_buy, "reinvest_amount"].sum()
            outflow  = closed.loc[mask_sell, "sell_value"].sum() if not closed.empty else 0
            monthly_rows.append({
                "YearMonth":             str(p),
                "SIP Invested (Bank)":   round(sip, 2),
                "Reinvested Amount":     round(reinvest, 2),
                "Inflow (Invested)":     round(sip + reinvest, 2),
                "Outflow (Sell Value)":  round(outflow, 2),
                "Net Cash Flow":         round(outflow - sip - reinvest, 2),
            })
    monthly_df = pd.DataFrame(monthly_rows)

    # ── Yearly Summary ────────────────────────────────────────────────────
    yearly_rows = []
    if not trades_df.empty:
        trades_df["FY"] = trades_df["buy_date"].apply(
            lambda d: d.year if pd.notna(d) and d.month > 3 else (d.year - 1 if pd.notna(d) else 0)
        )
        if not closed.empty and "FY" not in closed.columns:
            closed["FY"] = closed["sell_date"].apply(
                lambda d: d.year if pd.notna(d) and d.month > 3 else (d.year - 1 if pd.notna(d) else 0)
            )
        for fy in sorted(trades_df["FY"].unique()):
            mask_fy_buy  = trades_df["FY"] == fy
            mask_fy_sell = (closed["FY"] == fy) if (not closed.empty and "FY" in closed.columns) else pd.Series(False, index=closed.index)
            sip      = trades_df.loc[mask_fy_buy, "sip_amount"].sum()
            reinvest = trades_df.loc[mask_fy_buy, "reinvest_amount"].sum()
            outflow  = closed.loc[mask_fy_sell, "sell_value"].sum() if not closed.empty else 0
            profit   = closed.loc[mask_fy_sell, "gain"].sum() if not closed.empty else 0
            tax_paid = tax_df.loc[tax_df["FY"] == fy, "Tax Paid"].sum() if not tax_df.empty else 0
            yearly_rows.append({
                "FY":                    fy,
                "SIP Invested (Bank)":   round(sip, 2),
                "Reinvested Amount":     round(reinvest, 2),
                "Inflow (Invested)":     round(sip + reinvest, 2),
                "Outflow (Sell Value)":  round(outflow, 2),
                "Total Realized Profit": round(profit, 2),
                "Tax Paid":              round(tax_paid, 2),
                "Net Cash Flow":         round(outflow - sip - reinvest, 2),
            })
    yearly_df = pd.DataFrame(yearly_rows)

    return {
        "trades":          trades_df,
        "closed":          closed,
        "open":            open_t,
        "tax_df":          tax_df,
        "monthly_df":      monthly_df,
        "yearly_df":       yearly_df,
        "pre_tax_xirr":    pre_tax_xirr,
        "after_tax_xirr":  after_tax_xirr,
        "terminal_value":  terminal,
    }

# ─── EXCEL EXPORT ─────────────────────────────────────────────────────────────
def export_to_excel(analytics):
    output = io.BytesIO()
    summary_df = pd.DataFrame([{
        "Total Trades":      len(analytics["trades"]),
        "Open Trades":       len(analytics["open"]),
        "Closed Trades":     len(analytics["closed"]),
        "Terminal Value":    round(analytics["terminal_value"], 2),
        "Pre-Tax XIRR":      f"{analytics['pre_tax_xirr']:.2%}",
        "After-Tax XIRR":    f"{analytics['after_tax_xirr']:.2%}",
        "Total Tax Paid":    round(analytics["tax_df"]["Tax Paid"].sum(), 2) if not analytics["tax_df"].empty else 0,
    }])
    sheets = {
        "Summary":         summary_df,
        "All_Trades":      analytics["trades"],
        "Open_Trades":     analytics["open"],
        "Closed_Trades":   analytics["closed"],
        "Tax_Summary":     analytics["tax_df"],
        "Monthly_Summary": analytics["monthly_df"],
        "Yearly_Summary":  analytics["yearly_df"],
    }
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            if df is not None and not df.empty:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                ws = writer.sheets[sheet_name]
                ws.freeze_panes = "A2"
                for cell in ws[1]:
                    cell.font      = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center")
                for col in ws.columns:
                    max_len = max(len(str(c.value)) if c.value else 0 for c in col)
                    ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 4
    output.seek(0)
    return output

# ════════════════════════════════════════════════════════════════════════════
# UI
# ════════════════════════════════════════════════════════════════════════════
def main():
    st.title("📈 Mutual Fund Investment Tracker")

    page = st.sidebar.radio("Navigate", [
        "🏠 Dashboard",
        "➕ Add Fund",
        "💰 Add Trade",
        "🔔 Target Alerts & Sell",
        "📊 Analytics",
        "📥 Export to Excel",
    ])

    # ══════════════════════════════════════════
    # DASHBOARD
    # ══════════════════════════════════════════
    if page == "🏠 Dashboard":
        st.header("Portfolio Dashboard")
        funds_df  = read_sheet(SHEET_FUNDS,  FUNDS_COLS)
        trades_df = read_sheet(SHEET_TRADES, TRADES_COLS)

        if funds_df.empty:
            st.info("No funds added yet. Go to ➕ Add Fund to get started.")
            return

        num_cols = ["remaining_units", "buy_nav", "target_nav",
                    "invested_amount", "sip_amount", "reinvest_amount"]
        for col in num_cols:
            if col in trades_df.columns:
                trades_df[col] = pd.to_numeric(trades_df[col], errors="coerce").fillna(0)

        for _, fund in funds_df.iterrows():
            fund_id  = fund["fund_id"]
            amfi     = fund["amfi_code"]
            _, nav, nav_date = fetch_nav(amfi)

            ft   = trades_df[trades_df["fund_id"] == fund_id] if not trades_df.empty else pd.DataFrame()
            open_t  = ft[ft["status"] == "Open"]   if not ft.empty else pd.DataFrame()
            closed_t = ft[ft["status"] == "Closed"] if not ft.empty else pd.DataFrame()

            total_invested  = ft["invested_amount"].sum() if not ft.empty else 0
            current_value   = open_t["remaining_units"].sum() * nav if (not open_t.empty and nav) else 0
            cost_of_open    = (open_t["remaining_units"] * open_t["buy_nav"]).sum() if not open_t.empty else 0
            unrealized_gain = current_value - cost_of_open

            with st.expander(f"📦 {fund['fund_name']}", expanded=True):
                c1, c2, c3, c4, c5 = st.columns(5)
                c1.metric("Current NAV",    f"₹{nav:,.2f}" if nav else "N/A", nav_date)
                c2.metric("Total Invested", f"₹{total_invested:,.0f}")
                c3.metric("Current Value",  f"₹{current_value:,.0f}")
                c4.metric("Unrealized Gain",f"₹{unrealized_gain:,.0f}",
                          f"{(unrealized_gain/cost_of_open*100):.1f}%" if cost_of_open else "0%")
                c5.metric("Open / Closed",  f"{len(open_t)} / {len(closed_t)}")

                if not open_t.empty and nav:
                    hitting = open_t[pd.to_numeric(open_t["target_nav"], errors="coerce") <= nav]
                    if not hitting.empty:
                        st.warning(f"🎯 {len(hitting)} lot(s) have hit their Target NAV! Go to Target Alerts.")

    # ══════════════════════════════════════════
    # ADD FUND
    # ══════════════════════════════════════════
# Date and NAV fetch OUTSIDE the form so it reacts to date changes
        c1, c2 = st.columns(2)
        trade_date = c1.date_input("Trade Date", value=date.today())
        use_live   = c1.checkbox("Use Live NAV", value=True)

        _, date_nav, date_nav_date = fetch_nav(fund_row["amfi_code"], for_date=trade_date)

        if use_live:
            if date_nav:
                c2.info(f"NAV on {trade_date}: ₹{date_nav} ({date_nav_date})")
                nav_val = date_nav
            else:
                c2.warning("NAV not available for this date.")
                nav_val = None
        else:
            nav_val = c2.number_input(
                "Enter NAV manually",
                value=float(date_nav) if date_nav else 0.01,
                min_value=0.01,
                step=0.01
            )

        with st.form("add_trade"):
            sip_amount      = st.number_input("SIP Amount (from your Bank) ₹", min_value=0.0, step=500.0, value=15000.0)
            reinvest_amount = st.number_input("Reinvested Amount ₹ (0 if fresh SIP)", min_value=0.0, step=100.0, value=0.0)

            total_invest = sip_amount + reinvest_amount
            units        = round(total_invest / nav_val, 6) if nav_val else 0
            target_nav   = round(nav_val * (1 + float(fund_row["target_pct"])), 4) if nav_val else 0

            st.info(f"**Units:** {units:.4f} | **Target NAV:** ₹{target_nav:.2f} | **Total Invested:** ₹{total_invest:,.2f}")

            if st.form_submit_button("✅ Add Trade"):
                if nav_val and total_invest > 0:
                    append_row(SHEET_TRADES, {
                        "trade_id":        str(uuid.uuid4())[:12],
                        "fund_id":         fund_id,
                        "fund_name":       fund_name_sel,
                        "buy_date":        str(trade_date),
                        "buy_nav":         round(nav_val, 4),
                        "units":           units,
                        "remaining_units": units,
                        "invested_amount": round(total_invest, 2),
                        "sip_amount":      round(sip_amount, 2),
                        "reinvest_amount": round(reinvest_amount, 2),
                        "target_nav":      target_nav,
                        "status":          "Open",
                        "sell_date":       "",
                        "sell_nav":        "",
                        "sell_value":      "",
                        "gain":            "",
                        "holding_days":    "",
                        "tax_type":        "",
                    }, TRADES_COLS)
                    st.success(f"✅ Trade added! {units:.4f} units @ ₹{nav_val} | Target: ₹{target_nav:.2f}")
                    st.cache_data.clear()

    # ══════════════════════════════════════════
    # TARGET ALERTS & SELL
    # ══════════════════════════════════════════
    elif page == "🔔 Target Alerts & Sell":
        st.header("🔔 Target NAV Alerts & Sell")

        funds_df  = read_sheet(SHEET_FUNDS,  FUNDS_COLS)
        trades_df = read_sheet(SHEET_TRADES, TRADES_COLS)

        if trades_df.empty:
            st.info("No trades found.")
            return

        num_cols = ["buy_nav", "remaining_units", "target_nav", "units",
                    "sip_amount", "reinvest_amount"]
        for col in num_cols:
            trades_df[col] = pd.to_numeric(trades_df[col], errors="coerce").fillna(0)
        trades_df["buy_date"] = pd.to_datetime(trades_df["buy_date"], errors="coerce")

        open_trades = trades_df[trades_df["status"] == "Open"].copy()
        if open_trades.empty:
            st.success("No open trades.")
            return

        # Fetch live NAV per fund
        nav_map = {}
        for _, fund in funds_df.iterrows():
            _, nav, _ = fetch_nav(fund["amfi_code"])
            nav_map[fund["fund_id"]] = nav or 0

        open_trades["current_nav"]    = open_trades["fund_id"].map(nav_map)
        open_trades["target_reached"] = open_trades["current_nav"] >= open_trades["target_nav"]
        open_trades["gain_pct"]       = (
            (open_trades["current_nav"] - open_trades["buy_nav"]) / open_trades["buy_nav"]
        )

        hitting     = open_trades[open_trades["target_reached"]]
        not_hitting = open_trades[~open_trades["target_reached"]]

        # ── Lots that hit target ──────────────────────────────────────────
        if not hitting.empty:
            st.error(f"🎯 {len(hitting)} lot(s) have reached their Target NAV!")

            for _, lot in hitting.iterrows():
                with st.expander(
                    f"🎯 {lot['fund_name']} | Bought {str(lot['buy_date'])[:10]} "
                    f"@ ₹{lot['buy_nav']:.2f} | Target ₹{lot['target_nav']:.2f} "
                    f"| Now ₹{lot['current_nav']:.2f} (+{lot['gain_pct']:.1%})",
                    expanded=True
                ):
                    c1, c2, c3 = st.columns(3)
                    c1.metric("Remaining Units", f"{lot['remaining_units']:.4f}")
                    c2.metric("Current Value",   f"₹{lot['remaining_units'] * lot['current_nav']:,.2f}")
                    c3.metric("Gain",            f"{lot['gain_pct']:.1%}")

                    st.markdown("### 🤔 Did you sell this lot?")

                    with st.form(f"sell_{lot['trade_id']}"):
                        sc1, sc2 = st.columns(2)
                        sell_date     = sc1.date_input("Sell Date", value=date.today(),
                                                        key=f"sd_{lot['trade_id']}")
                        use_live_sell = sc1.checkbox("Use Live NAV for sell",
                                                      value=True, key=f"lv_{lot['trade_id']}")
                        if use_live_sell:
                            sell_nav = lot["current_nav"]
                            sc2.info(f"Sell NAV: ₹{sell_nav:.2f}")
                        else:
                            sell_nav = sc2.number_input("Sell NAV",
                                value=float(lot["current_nav"]),
                                key=f"snv_{lot['trade_id']}")

                        units_to_sell = st.number_input(
                            "Units to Sell (FIFO from oldest lot)",
                            value=float(lot["remaining_units"]),
                            max_value=float(lot["remaining_units"]),
                            step=0.001,
                            key=f"us_{lot['trade_id']}"
                        )

                        c_yes, c_no = st.columns(2)
                        yes_btn = c_yes.form_submit_button("✅ YES — I Sold", type="primary")
                        no_btn  = c_no.form_submit_button("❌ NO — Not Yet")

                        if yes_btn:
                            sell_records, err = process_fifo_sell(
                                lot["fund_id"], sell_date, sell_nav, units_to_sell
                            )
                            if err:
                                st.error(err)
                            else:
                                with st.spinner("Recording sale (FIFO)..."):
                                    commit_fifo_sell(sell_records, sell_date, sell_nav)
                                total_gain = sum(r["gain"] for r in sell_records)
                                st.success(
                                    f"✅ Sold {units_to_sell:.4f} units @ ₹{sell_nav:.2f} | "
                                    f"Gain: ₹{total_gain:,.2f} | "
                                    f"Tax: {'LTCG' if sell_records[0]['tax_type'] == 'LTCG' else 'STCG'}"
                                )
                                st.cache_data.clear()
                                st.rerun()

                        if no_btn:
                            st.info("Lot kept open.")

        # ── Lots awaiting target ──────────────────────────────────────────
        if not not_hitting.empty:
            st.subheader("📋 Lots Awaiting Target NAV")
            display = not_hitting[[
                "fund_name", "buy_date", "buy_nav", "target_nav",
                "current_nav", "remaining_units", "gain_pct"
            ]].copy()
            display.columns = [
                "Fund", "Buy Date", "Buy NAV", "Target NAV",
                "Current NAV", "Units", "Gain %"
            ]
            display["Buy Date"]    = display["Buy Date"].dt.date
            display["Gain %"]      = display["Gain %"].map("{:.1%}".format)
            display["Buy NAV"]     = display["Buy NAV"].map("₹{:.2f}".format)
            display["Target NAV"]  = display["Target NAV"].map("₹{:.2f}".format)
            display["Current NAV"] = display["Current NAV"].map("₹{:.2f}".format)
            st.dataframe(display, use_container_width=True)

    # ══════════════════════════════════════════
    # ANALYTICS
    # ══════════════════════════════════════════
    elif page == "📊 Analytics":
        st.header("📊 Portfolio Analytics")

        funds_df = read_sheet(SHEET_FUNDS, FUNDS_COLS)
        if funds_df.empty:
            st.info("No data yet.")
            return

        fund_options = {"All Funds": None}
        fund_options.update(dict(zip(funds_df["fund_name"], funds_df["fund_id"])))
        sel_name = st.selectbox("Filter by Fund", list(fund_options.keys()))
        sel_id   = fund_options[sel_name]

        with st.spinner("Computing analytics..."):
            a = compute_analytics(sel_id)

        if not a:
            st.info("No trades to analyze.")
            return

        # Summary
        st.subheader("Summary")
        c1, c2, c3, c4, c5, c6 = st.columns(6)
        c1.metric("Total Trades",     len(a["trades"]))
        c2.metric("Open Trades",      len(a["open"]))
        c3.metric("Closed Trades",    len(a["closed"]))
        c4.metric("Terminal Value",   f"₹{a['terminal_value']:,.0f}")
        c5.metric("Pre-Tax XIRR",     f"{a['pre_tax_xirr']:.2%}")
        c6.metric("After-Tax XIRR",   f"{a['after_tax_xirr']:.2%}")

        if not a["tax_df"].empty:
            st.metric("Total Tax Paid", f"₹{a['tax_df']['Tax Paid'].sum():,.2f}")

        tab1, tab2, tab3, tab4, tab5 = st.tabs([
            "📅 Monthly", "📆 Yearly", "🧾 Tax", "📂 Open Trades", "✅ Closed Trades"
        ])
        with tab1:
            if not a["monthly_df"].empty:
                st.dataframe(a["monthly_df"], use_container_width=True)
        with tab2:
            if not a["yearly_df"].empty:
                st.dataframe(a["yearly_df"], use_container_width=True)
        with tab3:
            if not a["tax_df"].empty:
                st.dataframe(a["tax_df"], use_container_width=True)
        with tab4:
            if not a["open"].empty:
                st.dataframe(a["open"], use_container_width=True)
        with tab5:
            if not a["closed"].empty:
                st.dataframe(a["closed"], use_container_width=True)

    # ══════════════════════════════════════════
    # EXPORT
    # ══════════════════════════════════════════
    elif page == "📥 Export to Excel":
        st.header("📥 Export to Excel")
        with st.spinner("Preparing..."):
            a = compute_analytics()
        if not a:
            st.info("No data to export.")
            return
        excel_data = export_to_excel(a)
        st.download_button(
            label="📥 Download Excel Report",
            data=excel_data,
            file_name=f"MF_Tracker_{date.today()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


if __name__ == "__main__":
    main()
